from .base import *
import stripe
from django.conf import settings
from django.shortcuts import render
from django.views.generic.detail import DetailView
from django.views.generic.list import ListView
from django.views.generic.edit import CreateView, UpdateView, FormView
from django.urls import reverse
from django.http import Http404, HttpResponseRedirect, HttpResponse
from django.contrib import messages
from django.db.models import Q
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from houses.mixins import HouseAccountMixin
from events.models import Event, Attendee, Answer, AttendeeCommonQuestions, EventQuestion

from events.forms import AttendeeForm


# Get Event Function
def get_event(slug):
	try:
		event = Event.objects.get(slug=slug)
	except Exception as e:
		print(e)
		raise Http404
	return event


# Create your views here.
class AttendeeListView(HouseAccountMixin, ListView):
	model = Attendee
	template_name = "events/attendees/event_attendees.html"

	def get(self, request, *args, **kwargs):
		return render(request, self.template_name, self.get_context_data())

	def get_event(self, slug):
		try:
			event = Event.objects.get(slug=slug)
			return event
		except Exception as e:
			print(e)
			raise Http404


	def export_to_excel(self, event, attendees):

		event_questions = EventQuestion.objects.filter(event=event, order_question=False)

		response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
		response['Content-Disposition'] = 'attachment; filename={event}-attendees.xlsx'.format(event=event.slug)
		workbook = Workbook()

		# Get active worksheet/tab
		worksheet = workbook.active
		worksheet.title = 'Attendees'

		# Define the titles for columns
		columns = ['Name', 'Ticket', 'Gender', 'Age', 'Refunded', 'Note']
		for event_question in event_questions:
			columns.append(event_question.question.title)
		row_num = 1

		# Assign the titles for each cell of the header
		for col_num, column_title in enumerate(columns, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = column_title

		# Iterate through all movies
		for attendee in attendees:
			row_num += 1

			answers = Answer.objects.filter(attendee=attendee)
			
			# Define the data for each cell in the row 
			row = []
			row.append(attendee.name)
			row.append(attendee.ticket.title)
			if attendee.gender:
				row.append(attendee.gender)
			else:
				row.append("N/A")
			
			if attendee.age:
				row.append(attendee.age)
			else:
				row.append("N/A")
			
			if attendee.active:
				row.append("No")
			else:
				row.append("Yes")

			if attendee.note:
				row.append(attendee.note)
			else:
				row.append("No Notes")

			for event_question in event_questions:
				value = "N/A"
				for answer in answers:
					if answer.question == event_question:
						value = answer.value
					else:
						value = "N/A"
				row.append(value)
		

			# Assign the data for each cell of the row 
			for col_num, cell_value in enumerate(row, 1):
				wrapped_alignment = Alignment(vertical='center', wrap_text=True)
				cell = worksheet.cell(row=row_num, column=col_num)
				cell.value = cell_value
				cell.alignment = wrapped_alignment


		# Change column widths
		name_column_letter = get_column_letter(1)
		ticket_name_column_letter = get_column_letter(2)
		note_column_letter = get_column_letter(6)

		name_column_dimensions = worksheet.column_dimensions[name_column_letter]
		ticket_name_column_dimensions = worksheet.column_dimensions[ticket_name_column_letter]
		note_column_dimensions = worksheet.column_dimensions[note_column_letter]

		name_column_dimensions.width = 30
		ticket_name_column_dimensions.width = 30
		note_column_dimensions.width = 50

		for x in range(7, (event_questions.count() + 7)):
			column_letter = get_column_letter(x)
			column_dimensions = worksheet.column_dimensions[column_letter]
			column_dimensions.width = 40

		workbook.save(response)

		return response

	def post(self, request, *args, **kwargs):
		data = request.POST
		print(data)
		event = self.get_event(self.kwargs['slug'])
		all_attendees = Attendee.objects.filter(order__event=event).order_by('order__created_at')

		if 'Export To Excel' in data:
			return self.export_to_excel(event, all_attendees)

		search_terms = data["search"].split()

		if data["search"] == '':
			attendees = all_attendees
		else:
			counter = 0
			for search_term in search_terms:
				if counter == 0:
					attendees = all_attendees.filter(Q(name__icontains=search_term) | Q(ticket__title__icontains=search_term))
				else:
					attendees = attendees.filter(Q(name__icontains=search_term) | Q(ticket__title__icontains=search_term))
				print(counter)
				counter += 1
		
		attendees = attendees[:100]
		print(attendees)
		html = render_to_string('events/attendees/attendees-dynamic-table-body.html', {'attendees': attendees, 'request':request})
		return HttpResponse(html)

	def get_context_data(self, *args, **kwargs):
		context = {}
		house = self.get_house()
		event = get_event(self.kwargs['slug'])
		attendees = Attendee.objects.filter(order__event=event, order__failed=False)
		print(attendees)
		
		context["house"] = house
		context["attendees"] = attendees
		context["event"] = event
		context["event_tab"] = True
		context["dashboard_events"] = self.get_events()
		return context




class AttendeeDetailView(HouseAccountMixin, FormView):
	model = Attendee
	template_name = "events/attendees/event_attendees_detail.html"

	def get_success_url(self):
		view_name = "events:attendee_detail"
		return reverse(view_name, kwargs={"slug": self.kwargs['slug'], "attendee_id": self.kwargs['attendee_id']})

	def get_attendee(self, attende_id):
		try:
			attendee = Attendee.objects.get(id=attende_id)
			return attendee
		except Exception as e:
			print(e)
			raise Http404

	def get_context_data(self, *args, **kwargs):
		context = {}

		house = self.get_house()
		slug = self.kwargs['slug']
		attendee_id = self.kwargs['attendee_id']
		event = get_event(slug)
		attendee = self.get_attendee(attendee_id)
		answers = Answer.objects.filter(attendee=attendee)
		attendee_common_questions = AttendeeCommonQuestions.objects.get(event=event)
		print(attendee_common_questions.age)

		context["house"] = house
		context["attendee_common_questions"] = attendee_common_questions
		context["event"] = event
		context["attendee"] = attendee
		context["answers"] = answers
		context["event_tab"] = True
		context["dashboard_events"] = self.get_events()
		return context

	def get(self, request, *args, **kwargs):
		return self.render_to_response(self.get_context_data())

	def post(self, request, *args, **kwargs):
		data = request.POST
		slug = kwargs['slug']
		attendee_id = kwargs['attendee_id']
		event = get_event(slug)
		attendee = self.get_attendee(attendee_id)
		answers = Answer.objects.filter(attendee=attendee)

		if 'name' in data:
			attendee.name = data['name']
			attendee.save()

		if 'email' in data:
			attendee.email = data['email']
			attendee.save()

		if 'gender' in data:
			attendee.gender = data['gender']
			attendee.save()

		if 'age' in data:
			attendee.age = data['age']
			attendee.save()

		if 'note' in data:
			attendee.note = data['note']
			attendee.save()

		for answer in answers:
			value = data["%s_%s" % (answer.question.pk, answer.attendee.id)]
			answer.value = value
			answer.save()

		messages.success(request, 'Attendee Updated')

		return self.render_to_response(self.get_context_data())







