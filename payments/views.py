from django.shortcuts import render
import decimal
import stripe 
import datetime
from django.db.models import Sum
from django.shortcuts import render
from django.urls import reverse
from django.views import View
from django.views.generic.edit import CreateView, UpdateView, FormView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.contrib import messages
from itertools import chain
from operator import attrgetter
from django.conf import settings
from django.http import Http404, HttpResponseRedirect, HttpResponse
from django.utils import timezone

from django.core import mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags

from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from twilio.rest import Client
from twilio.rest import Client

from weasyprint import HTML, CSS
from django.db.models import Sum

from houses.mixins import HouseAccountMixin
from houses.models import HouseUser
from donations.models import Donation
from profiles.models import Profile
from .models import Payout, Transaction, Refund, HousePayment, HouseBalance, HouseBalanceLog, PayoutSetting
from events.models import EventRefundRequest, EventOrder, EventOrderRefund
from .forms import AddFundsForm, PayoutForm, AddBankTransferForm


# Create your views here.

class InvoiceView(HouseAccountMixin, View):
	template_name = "payments/invoice.html"

	def view_pdf_statement(self, pdf_context):

		# PDF Attachment
		pdf_content = render_to_string('pdfs/statement.html', pdf_context)
		pdf_css = CSS(string=render_to_string('pdfs/statement.css'))

		# Creating http response
		response = HttpResponse(content_type='application/pdf;')
		response['Content-Disposition'] = 'inline; filename=tickets.pdf'

		pdf_file = HTML(string=pdf_content).write_pdf(
			response, stylesheets=[pdf_css])
		return response


	def export_to_excel(self, house, year, month, house_balance_logs_month):

		response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
		response['Content-Disposition'] = 'attachment; filename={house}_statement_{year}_{month}.xlsx'.format(house=house.slug, year=year, month=month)
		workbook = Workbook()

		# Get active worksheet/tab
		worksheet = workbook.active
		worksheet.title = f'{month} {year} Statement'

		TWO_PLACES = decimal.Decimal("0.01")

		# Define the titles for columns
		columns = ['Number', 'Date', 'Description', 'Name', 'Email', 'Phone', 'Address',
                    'Postal Code', 'Gross Amount', 'Fee', 'Net Amount', 'Tax - Donation Amount']
		row_num = 1

		# Assign the titles for each cell of the header
		for col_num, column_title in enumerate(columns, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell.value = column_title

		# Iterate through all house_balance_logs_month
		counter = 0
		for house_balance in house_balance_logs_month:
			row_num += 1

			# Define the data for each cell in the row
			row = []

			row.append(counter)

			row.append((house_balance.created_at).strftime("%d/%m/%Y"))

			if house_balance.transaction:
				if house_balance.transaction.donation_transaction:
					donation = Donation.objects.get(donation_type__house=house, transaction=house_balance.transaction)
					profile = Profile.objects.get(email=donation.email)
					row.append("Donation")
					row.append(donation.name)
					row.append(donation.email)
					if profile.phone:
						row.append(str(profile.phone))
					else:
						row.append("-----")

					if donation.address:
						row.append(donation.address)
					else:
						row.append("-----")
					
					if donation.postal_code:
						row.append(donation.postal_code)
					else:
						row.append("-----")
					
					
				else:
					event_order = EventOrder.objects.get(transaction=house_balance.transaction)
					profile = Profile.objects.get(email=event_order.email)
					row.append("Payment")
					row.append(event_order.name)
					row.append(event_order.email)
					if profile.phone:
						row.append(str(profile.phone))
					else:
						row.append("-----")

					row.append("-----")
					row.append("-----")

				# Gross, Fee and Net
				row.append((house_balance.transaction.amount).quantize(TWO_PLACES))
				total_fee = (house_balance.transaction.stripe_amount + house_balance.transaction.arqam_amount).quantize(TWO_PLACES)
				row.append(-total_fee)
				row.append((house_balance.transaction.house_amount).quantize(TWO_PLACES))

				# Donation Amount
				if house_balance.transaction.donation_transaction:
					donation = Donation.objects.get(donation_type__house=house, transaction=house_balance.transaction)
					row.append(donation.amount)
				else:
					row.append("-----")


			if house_balance.house_payment:
				row.append("Added Funds")
				row.append("-----")
				row.append("-----")
				row.append("-----")
				row.append("-----")
				row.append("-----")
				row.append((house_balance.house_payment.transaction.amount).quantize(TWO_PLACES))
				total_fee = (house_balance.house_payment.transaction.stripe_amount + house_balance.house_payment.transaction.arqam_amount).quantize(TWO_PLACES)
				row.append(-total_fee)
				row.append((house_balance.house_payment.transaction.house_amount).quantize(TWO_PLACES))
				row.append("-----")


			if house_balance.refund:
				event_order_refund = EventOrderRefund.objects.get(refund=refund)
				profile = Profile.objects.get(email=event_order_refund.order.email)
				row.append("Refund")
				row.append(event_order_refund.order.name)
				row.append(event_order_refund.order.email)

				if profile.phone:
					row.append(str(profile.phone))
				else:
					row.append("-----")
				
				row.append("-----")
				row.append("-----")

				row.append(-(house_balance.refund.amount).quantize(TWO_PLACES))
				row.append((house_balance.refund.fee()).quantize(TWO_PLACES))
				row.append(-(house_balance.refund.house_amount).quantize(TWO_PLACES))
				row.append("-----")


			if house_balance.payout:
				row.append("Payout")
				row.append("-----")
				row.append("-----")
				row.append("-----")
				row.append("-----")
				row.append("-----")
				row.append(-(house_balance.payout.amount).quantize(TWO_PLACES))
				row.append(0.00)
				row.append(-(house_balance.payout.amount).quantize(TWO_PLACES))
				row.append("-----")
			

			if house_balance.arqam_house_service_fee:
				if house_balance.arqam_house_service_fee.live_video:
					row.append("Virtual Event Fee")

				row.append("-----")
				row.append("-----")
				row.append("-----")
				row.append("-----")
				row.append("-----")

				if house_balance.arqam_house_service_fee.free:
					row.append(0.00)
				else:
					row.append((-house_balance.arqam_house_service_fee.amount).quantize(TWO_PLACES))

				row.append(0.00)
				
				if house_balance.arqam_house_service_fee.free:
					row.append(0.00)
				else:
					row.append((-house_balance.arqam_house_service_fee.amount).quantize(TWO_PLACES))

				row.append("-----")



			# Assign the data for each cell of the row
			for col_num, cell_value in enumerate(row, 1):
				wrapped_alignment = Alignment(vertical='center', wrap_text=True)
				cell = worksheet.cell(row=row_num, column=col_num)
				cell.value = cell_value
				cell.alignment = wrapped_alignment

			counter += 1

		# Change column widths
		for x in range(1, 13):
			column_letter = get_column_letter(x)
			column_dimensions = worksheet.column_dimensions[column_letter]
			column_dimensions.width = 25

		workbook.save(response)

		return response


	def get(self, request, *args, **kwargs):

		context = {}

		month_list = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                    'August', 'September', 'October', 'November', 'December']

		requested_year = self.kwargs["year"]
		requested_month = self.kwargs["month"]

		current_year = timezone.now().year
		current_month = timezone.now().month

		if requested_month not in month_list:
			raise Http404

		if requested_year < 2019 or requested_year > current_year:
			raise Http404


		house = self.get_house()
		house_balance = HouseBalance.objects.get(house=house)
		house_balance_logs = HouseBalanceLog.objects.filter(house_balance=house_balance)
		house_balance_logs_month = house_balance_logs.filter(created_at__month=month_list.index(requested_month)+1, created_at__year=requested_year)

		# ----------------- Payments
		payments = house_balance_logs_month.filter(transaction__isnull=False, transaction__donation_transaction=False)
		context["payments"] = payments
		if payments:
			payment_gross_amount = payments.aggregate(Sum('transaction__amount'))["transaction__amount__sum"]
			payment_total_amount = payments.aggregate(Sum('transaction__house_amount'))["transaction__house_amount__sum"]
			payment_stripe_fees = payments.aggregate(Sum('transaction__stripe_amount'))["transaction__stripe_amount__sum"]
			payment_arqam_fees = payments.aggregate(Sum('transaction__arqam_amount'))["transaction__arqam_amount__sum"]
			payment_total_fees = payment_stripe_fees + payment_arqam_fees
		else:
			payment_gross_amount = 0
			payment_total_amount = 0
			payment_total_fees = 0

		context["payment_gross_amount"] = payment_gross_amount
		context["payment_total_amount"] = payment_total_amount
		context["payment_total_fees"] = payment_total_fees

		# ----------------- Donations
		donations = house_balance_logs_month.filter(transaction__isnull=False, transaction__donation_transaction=True)
		context["donations"] = donations
		if donations:
			donation_gross_amount = donations.aggregate(Sum('transaction__amount'))["transaction__amount__sum"]
			donations_total_amount = donations.aggregate(Sum('transaction__house_amount'))["transaction__house_amount__sum"]
			donation_stripe_fees = donations.aggregate(Sum('transaction__stripe_amount'))["transaction__stripe_amount__sum"]
			donation_arqam_fees = donations.aggregate(Sum('transaction__arqam_amount'))["transaction__arqam_amount__sum"]
			donation_total_fees = donation_stripe_fees + donation_arqam_fees
		else:
			donation_gross_amount = 0
			donations_total_amount = 0
			donation_total_fees = 0

		context["donation_gross_amount"] = donation_gross_amount
		context["donation_total_amount"] = donations_total_amount
		context["donation_total_fees"] = donation_total_fees

		# ------------------- Huouse Added Payments
		house_payments = house_balance_logs_month.filter(house_payment__isnull=False)
		context["house_payments"] = house_payments
		if house_payments:
			house_payments_gross_amount = house_payments.aggregate(Sum('house_payment__transaction__amount'))["house_payment__transaction__amount__sum"]
			house_payments_total_amount = house_payments.aggregate(Sum('house_payment__transaction__house_amount'))["house_payment__transaction__house_amount__sum"]
			house_payments_stripe_fees = house_payments.aggregate(Sum('house_payment__transaction__stripe_amount'))["house_payment__transaction__stripe_amount__sum"]
			house_payments_arqam_fees = house_payments.aggregate(Sum('house_payment__transaction__arqam_amount'))["house_payment__transaction__arqam_amount__sum"]
			house_payments_total_fees = house_payments_stripe_fees + house_payments_arqam_fees
		else:
			house_payments_gross_amount = 0
			house_payments_total_amount = 0
			house_payments_total_fees = 0

		context["house_payments_gross_amount"] = house_payments_gross_amount
		context["house_payments_total_amount"] = house_payments_total_amount
		context["house_payments_total_fees"] = house_payments_total_fees

		# ------------------- Refunded Payments
		refunds = house_balance_logs_month.filter(refund__isnull=False)
		context["refunds"] = refunds
		if refunds:
			refund_gross_amount = refunds.aggregate(Sum('refund__amount'))["refund__amount__sum"]
			refund_house_amount = refunds.aggregate(Sum('refund__house_amount'))["refund__house_amount__sum"]
			refund_total_fees = refund_gross_amount - refund_house_amount
			context["refund_gross_amount"] = refund_gross_amount
			context["refund_house_amount"] = refund_house_amount
			context["refund_total_fees"] = refund_total_fees
		else:
			refund_gross_amount = 0
			refund_house_amount = 0
			refund_total_fees = 0

		context["refund_gross_amount"] = refund_gross_amount
		context["refund_house_amount"] = refund_house_amount
		context["refund_total_fees"] = refund_total_fees


		# ------------------- Arqam House Services
		services = house_balance_logs_month.filter(arqam_house_service_fee__isnull=False)
		context["services"] = services
		service_gross_amount = 0
		for service in services:
			if not service.arqam_house_service_fee.free:
				service_gross_amount += service.arqam_house_service_fee.amount

		# if services:
		# 	service_gross_amount = services.aggregate(Sum('arqam_house_service_fee__amount'))["arqam_house_service_fee__amount__sum"]
		# else:
		# 	service_gross_amount = 0

		context["service_gross_amount"] = service_gross_amount

		# -------------------- Payouts
		payouts = house_balance_logs_month.filter(payout__isnull=False)
		context["payouts"] = payouts
		if payouts:
			payout_gross_amount = payouts.aggregate(Sum('payout__amount'))["payout__amount__sum"]
		else:
			payout_gross_amount = 0

		context["payout_gross_amount"] = payout_gross_amount


		# -------------------- Last Month Balance
		try:
			current_time = datetime.datetime(int(requested_year), int(month_list.index(requested_month)+1), 1)
			previous_month_last_house_balance = house_balance_logs.filter(created_at__lt=current_time).last()
			previous_month_last_house_balance = previous_month_last_house_balance.balance
		except Exception as e:
			print(e)
			previous_month_last_house_balance = 0

		gross_activity = payment_gross_amount + house_payments_gross_amount + donation_gross_amount - refund_gross_amount - service_gross_amount
		total_fees = payment_total_fees + house_payments_total_fees + service_gross_amount - refund_total_fees  + donation_total_fees
		net_activity = payment_total_amount + house_payments_total_amount + donations_total_amount - refund_house_amount - service_gross_amount
		end_of_month_balance = previous_month_last_house_balance + net_activity - payout_gross_amount

		context["previous_month_last_house_balance"] = previous_month_last_house_balance
		context["gross_activity"] = gross_activity
		context["total_fees"] = total_fees
		context["net_activity"] = net_activity
		context["end_of_month_balance"] = end_of_month_balance

		context["house"] = self.get_house()
		context["dashboard_events"] = self.get_events()
		context["year"] = requested_year
		context["month"] = requested_month
		context["statements_tab"] = True

		if 'pdf_statement' in self.request.GET:
			return self.view_pdf_statement(context)

		if 'export_to_excel' in self.request.GET:
			return self.export_to_excel(house, requested_year, requested_month, house_balance_logs_month)

		return render(request, self.template_name, context)



class InvoiceMonthView(HouseAccountMixin, View):
	template_name = "payments/choose_month.html"


	def get(self, request, *args, **kwargs):

		context = {}

		requested_year = self.kwargs["year"]


		month_list = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                    'August', 'September', 'October', 'November', 'December']
		
		house = self.get_house()
		house_balance = HouseBalance.objects.get(house=house)
		house_balance_log = HouseBalanceLog.objects.get(house_balance=house_balance, opening_balance=True)

		months = []
		current_year = timezone.now().year
		current_month = timezone.localtime(timezone.now()).month
		opening_balance_year = house_balance_log.created_at.year
		opening_balance_month = house_balance_log.created_at.month

		if requested_year < 2019 or requested_year > current_year:
			raise Http404

		# Scenario for same year and month as opening balance
		if current_year == opening_balance_year and current_month == opening_balance_month:
			print("Did it come here")
			print(current_month)
			return(HttpResponseRedirect(reverse('payments:invoice', kwargs={"year": current_year, "month": month_list[current_month-1]})))
			

		# Scenario for current year 
		elif requested_year == current_year:
			print("is it here?")
			for x in range(current_month):
				months.append(month_list[x])

		# Scenario for previous year 
		else:
			for x in range(12):
				months.append(month_list[x])
		
		context["year"] = requested_year
		context["house"] = self.get_house()
		context["dashboard_events"] = self.get_events()
		context["months"] = months
		context["statements_tab"] = True

		return render(request, self.template_name, context)




class InvoiceYearView(HouseAccountMixin, View):
	template_name = "payments/choose_year.html"


	def get(self, request, *args, **kwargs):

		context = {}

		house = self.get_house()
		house_balance = HouseBalance.objects.get(house=house)
		house_balance_log = HouseBalanceLog.objects.get(
			house_balance=house_balance, opening_balance=True)

		years = []
		current_year = timezone.now().year
		opening_balance_year = house_balance_log.created_at.year

		years_opened = current_year - opening_balance_year

		if current_year == opening_balance_year:
			# Navigate to months no need to show years
			return(HttpResponseRedirect(reverse('payments:invoice_month', kwargs={"year": current_year})))

		else:
			for x in range(0, years_opened + 1):
				print(x)
				years.append(opening_balance_year + x)

		context["house"] = self.get_house()
		context["dashboard_events"] = self.get_events()
		context["years"] = years
		context["statements_tab"] = True

		return render(request, self.template_name, context)





class UpdateBankTransferView(HouseAccountMixin, CreateView):
	template_name = "payments/add_bank.html"
	form_class = AddBankTransferForm
	model = PayoutSetting

	def get_success_url(self):
		view_name = "payments:payout_settings_list"
		return reverse(view_name)

	def get(self, request, *args, **kwargs):
		data = request.GET
		print(data)
		payment_settings_id = self.kwargs["id"]
		self.object = PayoutSetting.objects.get(id=payment_settings_id)
		return render(request, self.template_name, self.get_context_data())

	
	def post(self, request, *args, **kwargs):
		payment_settings_id = self.kwargs["id"]
		self.object = PayoutSetting.objects.get(id=payment_settings_id)
		data = request.POST
		house = self.get_house()
		form = self.get_form()
		if form.is_valid():
			return self.form_valid(form, request, data)
		else:
			return self.form_invalid(form)
	
	def form_valid(self, form, request, data):
		house = self.get_house()
		self.object = form.save()
		messages.success(request, 'Bank Account Updated!')
		valid_data = super(UpdateBankTransferView, self).form_valid(form)
		return valid_data

	def get_context_data(self, *args, **kwargs):
		context = {}
		house = self.get_house()
		form = self.get_form()
		payment_settings_id = self.kwargs["id"]
		payout_setting = PayoutSetting.objects.get(house=house, id=payment_settings_id)
		context["form"] = form
		context["payout_setting"] = payout_setting
		context["dashboard_events"] = self.get_events()
		context["house"] = self.get_house()
		context["update"] = True
		context["payout_methods_tab"] = True
		return context


	def form_invalid(self, form):
		print(form.errors)
		return self.render_to_response(self.get_context_data(form=form))



class AddBankTransferView(HouseAccountMixin, CreateView):
	template_name = "payments/add_bank.html"
	form_class = AddBankTransferForm
	model = PayoutSetting

	def get_success_url(self):
		view_name = "payments:payout_settings_list"
		return reverse(view_name)

	def get(self, request, *args, **kwargs):

		data = request.GET
		self.object = None
		return render(request, self.template_name, self.get_context_data())

	
	def post(self, request, *args, **kwargs):
		data = request.POST
		house = self.get_house()
		form = self.get_form()
		if form.is_valid():
			return self.form_valid(form, request)
		else:
			return self.form_invalid(form)
	
	def form_valid(self, form, request):
		house = self.get_house()
		form.instance.house = house
		self.object = form.save()
		messages.success(request, 'Payout Method Created!')
		valid_data = super(AddBankTransferView, self).form_valid(form)
		return valid_data

	def get_context_data(self, *args, **kwargs):
		context = {}
		house = self.get_house()
		form = self.get_form()
		context["form"] = form
		context["dashboard_events"] = self.get_events()
		context["house"] = self.get_house()
		context["payout_methods_tab"] = True
		return context


	def form_invalid(self, form):
		print(form.errors)
		return self.render_to_response(self.get_context_data(form=form))



class PayoutSettingsListView(HouseAccountMixin, View):
	template_name = "payments/payout_settings_list.html"

	def get_payout_settings(self, house):
		payout_settings = PayoutSetting.objects.filter(house=house)
		return payout_settings

	def get(self, request, *args, **kwargs):
		return render(request, self.template_name, self.get_context_data())

	def get_context_data(self, data=None, *args, **kwargs):
		context = {}
		house = self.get_house()
		payout_settings = self.get_payout_settings(house)

		context["payout_settings"] = payout_settings
		context["dashboard_events"] = self.get_events()
		context["house"] = self.get_house()
		context["payout_methods_tab"] = True
		return context




class PayoutView(HouseAccountMixin, FormView):
	template_name = "payments/payout.html"
	model = Payout

	def get_success_url(self):
		view_name = "payments:list"
		return reverse(view_name)

	def get(self, request, *args, **kwargs):
		house = self.get_house()
		house_balance = HouseBalance.objects.get(house=house)
		form = PayoutForm(house_balance.balance, house)
		return render(request, self.template_name, self.get_context_data(form=form))

	
	def post(self, request, *args, **kwargs):
		data = request.POST
		house = self.get_house()
		house_balance = HouseBalance.objects.get(house=house)
		form = PayoutForm(data=data, total=house_balance.balance, house=house)
		if form.is_valid():
			return self.form_valid(form, request, house_balance)
		else:
			return self.form_invalid(form)


	def send_payout_email_us(self, payout):
		subject = 'New payout for %s' % (payout.house.name)
		context = {}
		context["payout"] = payout
		context["house"] = payout.house
		context["payout_amount"] = '{0:.2f}'.format(payout.amount)
		html_message = render_to_string('emails/payout_notify_us.html', context)
		plain_message = strip_tags(html_message)
		from_email = 'Arqam House Payout <admin@arqamhouse.com>'
		to = ['info@arqamhouse.com', 'admin@arqamhouse.com']
		mail.send_mail(subject, plain_message, from_email, to, html_message=html_message)
		return "Done"

	def send_payout_email_them(self, payout):
		subject = 'Arqam House - Payout Confirmation'
		context = {}
		context["payout"] = payout
		context["house"] = payout.house
		context["payout_amount"] = '{0:.2f}'.format(payout.amount)
		html_message = render_to_string('emails/payout_notify_them.html', context)
		plain_message = strip_tags(html_message)
		from_email = 'Payout Information <info@arqamhouse.com>'
		to = [payout.house.email]
		mail.send_mail(subject, plain_message, from_email, to, html_message=html_message, fail_silently=True)
		return "Done"

	def send_text_message_us(self, payout):
		account_sid = settings.ACCOUNT_SID
		auth_token = settings.AUTH_TOKEN
		client = Client(account_sid, auth_token)
		message = client.messages.create(
                    body="%s has requested a payout. Payout Amount: $%s. \n- Arqam House" % (
                        payout.house.name, '{0:.2f}'.format(payout.amount)),
                    from_='+16475571902',
                    to='+16472985582'
                )

	
	def form_valid(self, form, request, house_balance):
		house = self.get_house()
		house_balance = HouseBalance.objects.get(house=house)
		total = decimal.Decimal('{0:.2f}'.format(house_balance.balance))
		amount = decimal.Decimal(form.cleaned_data["amount"])

		if amount > total:
			form.add_error("amount", "Please choose a number less than or equal to your total")
			return self.form_invalid(form)

		payout_setting = form.cleaned_data["payout_setting"]
		payout = Payout.objects.create(house=house, amount=amount, payout_setting=payout_setting)
		self.send_payout_email_us(payout)
		self.send_payout_email_them(payout)
		self.send_text_message_us(payout)
		messages.success(request, 'Payout Requested!')
		valid_data = super(PayoutView, self).form_valid(form)
		return valid_data



	def get_context_data(self, form, *args, **kwargs):
		context = {}
		house = self.get_house()
		house_balance = HouseBalance.objects.get(house=house)
		payout_settings = PayoutSetting.objects.filter(house=house)

		total = '{0:.2f}'.format(house_balance.balance)

		refund_requests = EventRefundRequest.objects.filter(
			order__event__house=house, processed=False, dismissed=False)
		context["refund_requests"] = refund_requests

		context["form"] = form
		context["payout_settings"] = payout_settings
		context["house_balance"] = house_balance
		context["total"] = total
		context["dashboard_events"] = self.get_events()
		context["house"] = self.get_house()
		context["payout_tab"] = True
		return context


	def form_invalid(self, form):
		print(form.errors)
		return self.render_to_response(self.get_context_data(form=form))




class PaymentListView(HouseAccountMixin, View):
	template_name = "payments/list.html"

	def get_house_balance_logs(self, house_balance):
		house_balance_logs = HouseBalanceLog.objects.filter(house_balance=house_balance).order_by("-created_at")
		return house_balance_logs

	def get(self, request, *args, **kwargs):
		return render(request, self.template_name, self.get_context_data())

	def post(self, request, *args, **kwargs):
		data = request.POST
		print(data)
		house = self.get_house()
		house_balance = HouseBalance.objects.get(house=house)
		house_balance_logs = self.get_house_balance_logs(house_balance)

		if not data['reset'] == 'reset':

			start = data["start"]
			end = data["end"]

			log_type = data["log_type"]
			print(log_type)

			if start and end:
				today = timezone.now()
				start_year = int(start[0:4])
				end_year = int(end[0:4])

				start_month = int(start[5:7])
				end_month = int(end[5:7])

				start_day = int(start[8:10])
				end_day = int(end[8:10])
				print(start_day)
				print(end_day)

				house_balance_logs = house_balance_logs.filter(created_at__gte=datetime.date(start_year, start_month, start_day), created_at__lte=datetime.date(end_year, end_month, end_day))

			if log_type:
				if log_type == 'transaction':
					house_balance_logs = house_balance_logs.filter(transaction__isnull=False)
				elif log_type == 'refund':
					house_balance_logs = house_balance_logs.filter(refund__isnull=False)
				elif log_type == 'payout':
					house_balance_logs = house_balance_logs.filter(payout__isnull=False)
				elif log_type == 'house_payment':
					house_balance_logs = house_balance_logs.filter(house_payment__isnull=False)
				else:
					house_balance_logs = house_balance_logs


		html = render_to_string('payments/house_balance_logs_dynamic_table_body.html', {'house_balance_logs': house_balance_logs, 'request': request})
		return HttpResponse(html)


	def get_context_data(self, data=None, *args, **kwargs):
		context = {}
		house = self.get_house()
		house_balance = HouseBalance.objects.get(house=house)
		house_balance_logs = self.get_house_balance_logs(house_balance)
		context["house_balance"] = house_balance
		context["house_balance_logs"] = house_balance_logs
		context["dashboard_events"] = self.get_events()
		context["house"] = self.get_house()
		context["payout_list"] = True
		return context



class AddFundsView(HouseAccountMixin, FormView):
	template_name = "payments/add_funds.html"
	form_class = AddFundsForm
	model = Transaction

	def get_success_url(self):
		view_name = "payments:list"
		return reverse(view_name)

	def get(self, request, *args, **kwargs):
		return render(request, self.template_name, self.get_context_data())


	def get_context_data(self, *args, **kwargs):
		context = {}
		form = self.get_form()
		context["public_key"] = settings.STRIPE_PUBLIC_KEY
		context["form"] = form
		context["house"] = self.get_house()
		context["dashboard_events"] = self.get_events()
		context["payout_tab"] = True
		return context

	
	def post(self, request, *args, **kwargs):
		data = request.POST
		form = self.get_form()
		if form.is_valid():
			return self.form_valid(form, request)
		else:
			return self.form_invalid(form)


	def form_valid(self, form, request):
		data = request.POST
		house = self.get_house()
		user = request.user
		print(data)
		stripe_token = data["stripeToken"]
		amount = decimal.Decimal(form.cleaned_data["amount"])

		stripe_charge_amount = amount * 100
		stripe_charge_amount = int(stripe_charge_amount)

		fee = (amount * decimal.Decimal(0.04)) + decimal.Decimal(0.30)
		stripe_amount = (amount * decimal.Decimal(0.029)) + decimal.Decimal(0.30)
		arqam_amount = fee - stripe_amount
		house_total = amount - fee
		
		try:
			stripe.api_key = settings.STRIPE_SECRET_KEY
			
			charge = stripe.Charge.create(
							amount = stripe_charge_amount,
							currency = 'cad',
							description = 'Funds added for %s' % (house.name),
							source = stripe_token,
							statement_descriptor = 'Arqam House Inc.',
						)
			print(charge)

			transaction = Transaction.objects.create(house=house,
			amount = amount,
			house_amount = house_total,
			stripe_amount = stripe_amount,
			arqam_amount = arqam_amount,
			house_payment = True,
			payment_id = charge['id'],
			last_four = charge.source['last4'],
			brand = charge.source['brand'],
			network_status = charge.outcome['network_status'],
			risk_level = charge.outcome['risk_level'],
			seller_message = charge.outcome['seller_message'],
			outcome_type = charge.outcome['type'],
			email = user.email,
			name = charge.source['name'],
			address_line_1 = charge.source['address_line1'],
			address_state = charge.source['address_state'],
			address_postal_code = charge.source['address_zip'],
			address_city = charge.source['address_city'],
			address_country = charge.source['address_country'],
			)
			house_payment = HousePayment.objects.create(transaction=transaction)
		except Exception as e:
			print(e)
			messages.success(request, 'An Error Occured. Please contact support.')
			return HttpResponseRedirect(self.get_success_url())
		

		messages.success(request, 'Funds Added!')
		valid_data = super(AddFundsView, self).form_valid(form)
		return valid_data

	def form_invalid(self, form):
		print(form.errors)
		return self.render_to_response(self.get_context_data(form=form))


